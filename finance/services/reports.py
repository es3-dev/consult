from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from django.contrib.auth.models import User
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


def filtered_movements(user: User, filters: dict) -> list[dict]:
    rows = []
    start = filters.get("start")
    end = filters.get("end")
    category = filters.get("category")
    kind = filters.get("type")

    if kind in ("", None, "income"):
        qs = user.incomes.select_related("category").all()
        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        if category:
            qs = qs.filter(category_id=category)
        rows.extend({"type": "Ingreso", "date": i.date, "category": i.category.name, "concept": i.source, "amount": i.amount} for i in qs)

    if kind in ("", None, "expense"):
        qs = user.expenses.select_related("category").all()
        if start:
            qs = qs.filter(date__gte=start)
        if end:
            qs = qs.filter(date__lte=end)
        if category:
            qs = qs.filter(category_id=category)
        rows.extend({"type": "Gasto", "date": e.date, "category": e.category.name, "concept": e.merchant, "amount": e.amount} for e in qs)

    return sorted(rows, key=lambda row: row["date"], reverse=True)


def report_summary(rows: list[dict]) -> dict:
    income = sum((row["amount"] for row in rows if row["type"] == "Ingreso"), Decimal("0.00"))
    expense = sum((row["amount"] for row in rows if row["type"] == "Gasto"), Decimal("0.00"))
    return {"income": income, "expense": expense, "balance": income - expense, "count": len(rows)}


def money(value: Decimal) -> str:
    return f"${float(value):,.0f}"


def excel_response(user: User, filters: dict) -> HttpResponse:
    rows = filtered_movements(user, filters)
    summary = report_summary(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte financiero"

    dark = "0F172A"
    soft = "F8FAFC"
    green = "DCFCE7"
    rose = "FFE4E6"

    ws.merge_cells("A1:E1")
    ws["A1"] = "Consult-App - Reporte financiero"
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=dark)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.append([])
    ws.append(["Generado", timezone.localtime().strftime("%Y-%m-%d %H:%M"), "", "Movimientos", summary["count"]])
    ws.append(["Ingresos", float(summary["income"]), "Gastos", float(summary["expense"]), "Balance"])
    ws.append(["", "", "", "", float(summary["balance"])])

    for cell in ("B4", "D4", "E5"):
        ws[cell].number_format = '$#,##0'
        ws[cell].font = Font(bold=True)

    ws.append([])
    header_row = ws.max_row + 1
    ws.append(["Tipo", "Fecha", "Categoria", "Concepto", "Monto"])
    for cell in ws[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        ws.append([row["type"], str(row["date"]), row["category"], row["concept"], float(row["amount"])])
        current = ws.max_row
        ws[f"E{current}"].number_format = '$#,##0'
        ws[f"A{current}"].fill = PatternFill("solid", fgColor=green if row["type"] == "Ingreso" else rose)

    for column, width in {"A": 14, "B": 14, "C": 22, "D": 36, "E": 16}.items():
        ws.column_dimensions[column].width = width
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center")

    ws.freeze_panes = f"A{header_row + 1}"
    ws.sheet_view.showGridLines = False

    stream = BytesIO()
    wb.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="consult-app-reporte.xlsx"'
    return response


def pdf_response(user: User, filters: dict) -> HttpResponse:
    rows = filtered_movements(user, filters)
    summary = report_summary(rows)
    stream = BytesIO()
    doc = SimpleDocTemplate(stream, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Title"], textColor=colors.HexColor("#0f172a"), fontSize=18, leading=22)
    small = ParagraphStyle("Small", parent=styles["BodyText"], textColor=colors.HexColor("#64748b"), fontSize=9)

    story = [
        Paragraph("Consult-App - Reporte financiero", title),
        Paragraph(f"Generado el {timezone.localtime().strftime('%Y-%m-%d %H:%M')}", small),
        Spacer(1, 0.18 * inch),
    ]

    summary_table = Table(
        [
            ["Ingresos", "Gastos", "Balance", "Movimientos"],
            [money(summary["income"]), money(summary["expense"]), money(summary["balance"]), str(summary["count"])],
        ],
        colWidths=[1.55 * inch, 1.55 * inch, 1.55 * inch, 1.55 * inch],
    )
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("BACKGROUND", (0, 1), (-1, 1), colors.HexColor("#f8fafc")),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 0.22 * inch)])

    data = [["Tipo", "Fecha", "Categoria", "Concepto", "Monto"]]
    for row in rows[:80]:
        data.append([row["type"], str(row["date"]), row["category"], row["concept"][:32], money(row["amount"])])

    table = Table(data, colWidths=[0.85 * inch, 0.95 * inch, 1.25 * inch, 2.4 * inch, 1.0 * inch], repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f172a")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (4, 1), (4, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#e2e8f0")),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(table)
    if len(rows) > 80:
        story.append(Spacer(1, 0.12 * inch))
        story.append(Paragraph("El PDF muestra los primeros 80 movimientos. Exporta Excel para ver el detalle completo.", small))

    doc.build(story)
    response = HttpResponse(stream.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="consult-app-reporte.pdf"'
    return response
